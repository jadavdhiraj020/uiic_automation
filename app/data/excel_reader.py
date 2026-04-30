"""
excel_reader.py
Reads the Excel file (.xls / .xlsx) from the scanned folder.

SEARCH STRATEGY (position-independent):
  1. Scan all cells in the sheet for the search_label text.
  2. From that label cell, look RIGHT across the same row for the first
     non-junk value (handles any column layout).
  3. If row_offset > 0, move down that many rows first, then scan right.
  4. col_offset is used ONLY as a tiebreaker hint — it no longer causes
     failures when the Excel layout shifts between claims.

This makes the reader robust regardless of where the label or value
appears in any given Excel file.
"""
import os
import re
import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, List, Optional, Tuple

logger = logging.getLogger(__name__)


# ── Excel library selection ───────────────────────────────────────────────────
def _open_workbook(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        return _XlrdWrapper(wb)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        return _OpenpyxlWrapper(wb)


class _XlrdWrapper:
    def __init__(self, wb):
        self._wb = wb

    def sheet_names(self):
        return self._wb.sheet_names()

    def get_sheet(self, name):
        try:
            sh = self._wb.sheet_by_name(name)
            return _XlrdSheetWrapper(sh)
        except Exception:
            return None

    def all_sheets(self):
        return [_XlrdSheetWrapper(self._wb.sheet_by_index(i))
                for i in range(self._wb.nsheets)]


class _XlrdSheetWrapper:
    def __init__(self, sh):
        self._sh = sh
        self.name = sh.name

    def rows(self):
        for r in range(self._sh.nrows):
            yield [self._sh.cell_value(r, c) for c in range(self._sh.ncols)]


class _OpenpyxlWrapper:
    def __init__(self, wb):
        self._wb = wb

    def sheet_names(self):
        return self._wb.sheetnames

    def get_sheet(self, name):
        if name in self._wb:
            return _OpenpyxlSheetWrapper(self._wb[name])
        return None

    def all_sheets(self):
        return [_OpenpyxlSheetWrapper(self._wb[n]) for n in self._wb.sheetnames]


class _OpenpyxlSheetWrapper:
    def __init__(self, sh):
        self._sh = sh
        self.name = sh.title

    def rows(self):
        for row in self._sh.iter_rows(values_only=True):
            yield [v if v is not None else "" for v in row]


# ── Junk values — never returned as a field value ─────────────────────────────
_JUNK_VALUES = {
    ':', 'rs', 'rs.', 'inr', '-', '--', 'n/a', 'nil', 'na',
    'attached', 'yes', 'no', 'period:', 'date:', 'amount',
    'estimated', 'assessed', 'particulars', 'description',
}

# Junk patterns — text that looks like a label, not a value
_JUNK_PATTERNS = [
    re.compile(r'^[a-z\s/&()%,.:]+$'),      # Pure text with no digits
    re.compile(r'^rs\.?\s*$', re.I),         # "Rs" or "Rs."
    re.compile(r'^\s*:\s*$'),                 # Just ":"
]


def _is_junk(val: Any) -> bool:
    """Return True if val should NOT be treated as a field value."""
    # 0 and 0.0 are valid numeric values — never junk
    if val == 0 or val == 0.0:
        return False
    if val in (None, ""):
        return True
    s = str(val).strip()
    if not s:
        return True
    sl = s.lower()
    if sl in _JUNK_VALUES:
        return True
    # Pure alphabetic strings with no digits are junk (they are labels)
    for pat in _JUNK_PATTERNS:
        if pat.match(sl) and not any(c.isdigit() for c in s):
            return True
    return False


# ── Core search — position-independent ───────────────────────────────────────

def _search_label(sheet, label: str, row_offset: int, col_offset: int,
                  is_date: bool = False,
                  allow_literal_values: bool = False,
                  allow_text_values: bool = False) -> Tuple[Optional[str], Optional[str]]:
    """
    Find label text anywhere in the sheet. Then:
      1. Move row_offset rows down.
      2. Try col_offset first (exact hint from config).
      3. If that's empty/junk, scan RIGHT from the label column to find
         the first non-junk value.
      4. This makes the search position-independent: the label can be in
         any column; the value just needs to be to the right on the same row.

    Returns tuple (cleaned string value, coordinate info string) or (None, None).
    """
    label_lower = " ".join(label.lower().split())  # normalize whitespace
    all_rows = list(sheet.rows())

    for r_idx, row in enumerate(all_rows):
        for c_idx, cell in enumerate(row):
            cell_str = " ".join(str(cell).strip().lower().split())  # collapse \n, \t, multi-space
            if cell_str and label_lower in cell_str:
                # Word-boundary check: prevent "TOTAL" matching "SUBTOTAL"
                idx = cell_str.find(label_lower)
                before_ok = (idx == 0) or not cell_str[idx - 1].isalnum()
                after_end = idx + len(label_lower)
                after_ok = (after_end >= len(cell_str)) or not cell_str[after_end].isalnum()
                if not (before_ok and after_ok):
                    continue
                # Found the label at (r_idx, c_idx)
                
                # ── Strategy 0: Inline value (e.g., "Mobile: 098761-35253" in one cell)
                inline_content = cell_str.replace(label_lower, "").strip(" :-\n\t")
                if len(inline_content) > 3:
                    orig_cell = str(cell)
                    # Try splitting by colon or just removing the label text
                    if ":" in orig_cell:
                        val = orig_cell.split(":", 1)[-1].strip(" -\n\t")
                    else:
                        val = re.sub(re.escape(label), "", orig_cell, flags=re.IGNORECASE).strip(" -:\n\t")
                    
                    if val:
                        result = _extract_value(val, is_date, allow_literal_values, allow_text_values)
                        if result is not None:
                            coord_str = f"R{r_idx+1}C{c_idx+1}"
                            logger.info(
                                f"  [{label}] found inline at {coord_str} = {result}"
                            )
                            return result, coord_str

                target_r = r_idx + row_offset

                if not (0 <= target_r < len(all_rows)):
                    continue

                target_row = all_rows[target_r]

                # ── Strategy 1: Try col_offset hint first ──────────────────
                hint_c = c_idx + col_offset
                if 0 <= hint_c < len(target_row):
                    val = target_row[hint_c]
                    result = _extract_value(val, is_date, allow_literal_values, allow_text_values)
                    if result is not None:
                        coord_str = f"R{target_r+1}C{hint_c+1}"
                        logger.info(
                            f"  [{label}] found at R{r_idx+1}C{c_idx+1}, "
                            f"value at {coord_str} = {result}"
                        )
                        return result, coord_str

                # ── Strategy 2: Scan right from label col to find first value
                for scan_c in range(c_idx + 1, len(target_row)):
                    val = target_row[scan_c]
                    result = _extract_value(val, is_date, allow_literal_values, allow_text_values)
                    if result is not None:
                        coord_str = f"R{target_r+1}C{scan_c+1}"
                        logger.info(
                            f"  [{label}] found at R{r_idx+1}C{c_idx+1}, "
                            f"value at {coord_str} (scan right) = {result}"
                        )
                        return result, coord_str

                # ── Strategy 3: REMOVED FOR SAFETY ─────────────────────────────
                # We no longer drop down to the next row automatically if row_offset=0.
                # If the value is not on the expected row, we stop to prevent grabbing wrong data.

    return None, None


def _extract_value(val: Any, is_date: bool,
                   allow_literal_values: bool = False,
                   allow_text_values: bool = False) -> Optional[str]:
    """
    Convert a raw cell value to a usable string.
    Returns None if the value is empty or junk.
    """
    if val == "" or val is None:
        return None
    if allow_literal_values:
        s = str(val).strip()
        if s and s.lower() in {"yes", "no"}:
            return s
    if allow_text_values:
        # Accept any non-empty text — skip the junk/pattern checks entirely
        # But still reject trivial separators like ":" or "-"
        s = str(val).strip().strip(":- ")
        return s if s else None
    if _is_junk(val):
        return None

    # Special case: val == 0 or 0.0 IS a valid value
    if val == 0 or val == 0.0:
        return "0"

    if is_date:
        date_result = _try_date_serial(val)
        if date_result:
            return date_result

    return _clean_value(val)


def _clean_value(val: Any) -> str:
    """Convert Excel cell value to a clean string."""
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return f"{val:.2f}"
    return str(val).strip()


def _initial_loss_75_percent(value: Any) -> str:
    """
    Use 75% of the Excel "initial loss assessment" amount.

    The portal ultimately wants rounded rupees, so keep UI preview and
    automation aligned by storing the rounded 75% amount in ClaimData.
    """
    raw = str(value).strip()
    if raw == "":
        return raw

    cleaned = raw.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if not match:
        return raw

    try:
        amount = Decimal(match.group(0))
    except (InvalidOperation, ValueError):
        logger.warning("  [MATH] Could not calculate 75%% initial loss from '%s'", value)
        return raw

    adjusted = (amount * Decimal("0.75")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return str(int(adjusted))


def _try_date_serial(val: Any) -> Optional[str]:
    """Try to convert xlrd float date serial to DD/MM/YYYY."""
    if not isinstance(val, float):
        return None
    # xlrd date serials for years 1990-2040 fall in ~32874-51544
    if 32874 < val < 51544:
        try:
            import xlrd
            t = xlrd.xldate_as_tuple(val, 0)
            if t[0] > 1990:
                return f"{t[2]:02d}/{t[1]:02d}/{t[0]}"
        except Exception:
            pass
    return None


def _format_date(raw: str) -> str:
    """Normalise any date string to DD/MM/YYYY for the portal."""
    if not raw:
        return ""
    if re.match(r"\d{2}/\d{2}/\d{4}", raw):
        return raw
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y",
                "%Y/%m/%d", "%B %d, %Y"):
        try:
            dt = datetime.strptime(raw.strip(), fmt)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            continue
    return raw


# ── Public API ────────────────────────────────────────────────────────────────

def read_excel(excel_path: str, config_dir: str):
    """
    Read Excel file and return a fully populated ClaimData instance.

    For each field in field_mapping.json:
      - Finds the label text anywhere in the configured sheet
      - Reads the value to the right of the label (position-independent)
      - Falls back to scanning right and then scanning below
      - Logs exactly which cell was read for each field
    """
    from app.data.data_model import ClaimData
    from app.utils import load_field_mapping

    # Use the user's custom field mapping from AppData if it exists,
    # otherwise fall back to the bundled default.
    mapping = load_field_mapping()

    wb = _open_workbook(excel_path)
    claim = ClaimData()

    found_count = 0
    missing_fields = []

    for field_name, cfg in mapping.items():
        if field_name.startswith("_"):
            continue

        if field_name == "surveyor_observation":
            claim.surveyor_observation = "ok"
            claim._excel_coords["surveyor_observation"] = "Fixed Value"
            claim._excel_logs.append("  📊 surveyor_observation: 'ok' (Source: Fixed Value)")
            logger.info("  [FIXED] surveyor_observation = ok")
            found_count += 1
            continue

        sheet_name = cfg.get("sheet", "ALL")
        
        labels = cfg.get("search_labels")
        if not labels:
            raw_label = cfg.get("search_label", "")
            if isinstance(raw_label, list):
                labels = raw_label
            else:
                labels = [raw_label] if raw_label else []

        row_off    = cfg.get("row_offset", 0)
        col_off    = cfg.get("col_offset", 1)
        is_date    = "date" in field_name
        allow_literal_values = bool(cfg.get("allow_literal_values"))
        allow_text_values    = bool(cfg.get("allow_text_values"))

        value = None
        found_label = ""

        for current_label in labels:
            if not current_label:
                continue
                
            if sheet_name == "ALL":
                for sh in wb.all_sheets():
                    value, coord = _search_label(
                        sh, current_label, row_off, col_off, is_date,
                        allow_literal_values, allow_text_values
                    )
                    if value:
                        sh_name = sh.name if hasattr(sh, 'name') else 'Sheet'
                        src_str = f"{coord} ({sh_name})"
                        claim._excel_coords[field_name] = src_str
                        claim._excel_logs.append(f"  📊 {field_name}: '{value}' (Source: {src_str})")
                        found_label = current_label
                        break
            else:
                sh = wb.get_sheet(sheet_name)
                if sh:
                    value, coord = _search_label(
                        sh, current_label, row_off, col_off, is_date,
                        allow_literal_values, allow_text_values
                    )
                    if value:
                        src_str = f"{coord} ({sheet_name})"
                        claim._excel_coords[field_name] = src_str
                        claim._excel_logs.append(f"  📊 {field_name}: '{value}' (Source: {src_str})")
                        found_label = current_label
                else:
                    if current_label == labels[-1]: # Only warn on the last fallback try
                        logger.warning(f"  [{field_name}] Sheet '{sheet_name}' not found in workbook")

            if value:
                break  # Found it, stop trying fallback labels

        if value:
            if field_name == "date_of_survey":
                # ── Extract time from the date string OR adjacent cells ──────
                time_found = False

                # First: try parsing time from the value itself
                time_match = re.search(r"(\d{1,2})[.:]?(\d{2})?\s*([aA]\.?[mM]\.?|[pP]\.?[mM]\.?)", value)
                if time_match:
                    h = int(time_match.group(1))
                    m = time_match.group(2) or "00"
                    ampm = time_match.group(3).replace(".", "").lower()
                    if ampm == "pm" and h < 12:
                        h += 12
                    elif ampm == "am" and h == 12:
                        h = 0
                    claim.time_hh = f"{h:02d}"
                    claim.time_mm = f"{int(m):02d}"
                    time_found = True

                # Second: if no time in value, scan adjacent cells in the row
                if not time_found:
                    try:
                        all_sheets = wb.all_sheets() if sheet_name == "ALL" else [wb.get_sheet(sheet_name)]
                        for sh in all_sheets:
                            if sh is None:
                                continue
                            for r_idx, row in enumerate(sh.rows()):
                                for c_idx, cell in enumerate(row):
                                    cell_lower = str(cell).strip().lower()
                                    if found_label.lower() in cell_lower and cell_lower:
                                        target_r = r_idx + cfg.get("row_offset", 0)
                                        all_row_data = list(sh.rows())
                                        if 0 <= target_r < len(all_row_data):
                                            target_row = all_row_data[target_r]
                                            for tc in range(c_idx + 1, len(target_row)):
                                                tc_str = str(target_row[tc]).strip()
                                                tm = re.search(r"(\d{1,2})[.:]?(\d{2})?\s*([aA]\.?[mM]\.?|[pP]\.?[mM]\.?)", tc_str)
                                                if tm:
                                                    h = int(tm.group(1))
                                                    m = tm.group(2) or "00"
                                                    ampm = tm.group(3).replace(".", "").lower()
                                                    if ampm == "pm" and h < 12:
                                                        h += 12
                                                    elif ampm == "am" and h == 12:
                                                        h = 0
                                                    claim.time_hh = f"{h:02d}"
                                                    claim.time_mm = f"{int(m):02d}"
                                                    time_found = True
                                                    logger.info(f"  [TIME] Extracted from adjacent cell R{target_r+1}C{tc+1}: {claim.time_hh}:{claim.time_mm}")
                                                    break
                                                tm24 = re.search(r"\b([01]?\d|2[0-3]):([0-5]\d)\b", tc_str)
                                                if tm24:
                                                    claim.time_hh = f"{int(tm24.group(1)):02d}"
                                                    claim.time_mm = f"{int(tm24.group(2)):02d}"
                                                    time_found = True
                                                    logger.info(f"  [TIME] Extracted 24h from adjacent cell R{target_r+1}C{tc+1}: {claim.time_hh}:{claim.time_mm}")
                                                    break
                                            if time_found: break
                                    if time_found: break
                                if time_found: break
                            if time_found: break
                    except Exception as e:
                        logger.warning(f"  [TIME] Adjacent cell scan failed: {e}")

                if time_found:
                    logger.info(f"  [TIME] Survey time set: HH={claim.time_hh} MM={claim.time_mm}")

            if is_date:
                clean_date_val = re.sub(r"at.*$", "", str(value), flags=re.IGNORECASE).strip()
                value = _format_date(clean_date_val)

            if field_name == "initial_loss_amount":
                raw_initial_loss = value
                value = _initial_loss_75_percent(value)
                claim._excel_logs.append(
                    f"  📊 initial_loss_amount_75_percent: '{value}' "
                    f"(Source: 75% of Excel value '{raw_initial_loss}')"
                )
                logger.info(
                    "  [MATH] initial_loss_amount 75%% = %s (Excel value: %s)",
                    value,
                    raw_initial_loss,
                )

            setattr(claim, field_name, value)
            found_count += 1
            logger.info(f"  [FOUND] {field_name} = {value}")
        else:
            fallback = cfg.get("fallback_value")
            if fallback is not None:
                setattr(claim, field_name, fallback)
                logger.info(f"  [FALLBACK] {field_name} = {fallback}")
                claim._excel_logs.append(f"  📊 {field_name}: '{fallback}' (Source: Fallback Configuration)")
            else:
                labels_str = ' | '.join(labels)
                missing_fields.append(f"{field_name} (labels: '{labels_str}')")
                logger.warning(f"  [MISSING] {field_name}: labels '{labels_str}' not found or value empty")

    # Expected completion date no longer needs separate Excel extraction or +10 day calculation.
    # Keep it aligned with Date of Survey so every downstream consumer sees the same value.
    if claim.date_of_survey:
        claim.expected_completion_date = claim.date_of_survey
        if "expected_completion_date" not in claim._excel_coords:
            claim._excel_coords["expected_completion_date"] = claim._excel_coords.get("date_of_survey", "")
        logger.info(
            "  [SYNC] expected_completion_date = %s (same as date_of_survey)",
            claim.expected_completion_date,
        )

    # ── Calculate Total Claimed Amount ──────────────────────────────────────
    # To ensure UI preview and Backend automation are synchronized (Point 6),
    # we dynamically calculate the sum of surveyor charges here instead of
    # relying on the explicitly extracted field.
    try:
        calculated_total = sum(
            int(float(getattr(claim, k) or 0))
            for k in ["traveling_expenses", "professional_fee", "daily_allowance", "photo_charges"]
        )
        claim.total_claimed_amount = str(calculated_total)
        claim._excel_coords["total_claimed_amount"] = "Calculated"
        claim._excel_logs.append(f"  📊 total_claimed_amount: '{claim.total_claimed_amount}' (Source: Calculated)")
        logger.info(f"  [MATH] Calculated total_claimed_amount = {claim.total_claimed_amount}")
    except Exception as e:
        logger.warning(f"  [MATH] Failed to calculate total claimed amount: {e}")

    logger.info(f"Excel read complete: {found_count} fields found, "
                f"{len(missing_fields)} missing: {missing_fields}")

    # ── Payment Type Detection (keyword scan) ────────────────────────────────
    if not claim.payment_to:
        for sh in wb.all_sheets():
            sh_name = sh.name if hasattr(sh, 'name') else 'Sheet'
            for r_idx, row in enumerate(sh.rows()):
                for c_idx, cell in enumerate(row):
                    cell_text = " ".join(str(cell).strip().lower().split())
                    
                    if "payment to insured" in cell_text:
                        claim.payment_to = "INSURED"
                        src = f"R{r_idx+1}C{c_idx+1} ({sh_name})"
                        claim._excel_coords["payment_to"] = src
                        claim._excel_logs.append(f"  📊 payment_to: '{claim.payment_to}' (Source: {src})")
                        logger.info(f"  [FOUND] payment_to = {claim.payment_to} (keyword scan - payment to insured)")
                        break
                    elif "payment to repairer" in cell_text:
                        claim.payment_to = "REPAIRER"
                        src = f"R{r_idx+1}C{c_idx+1} ({sh_name})"
                        claim._excel_coords["payment_to"] = src
                        claim._excel_logs.append(f"  📊 payment_to: '{claim.payment_to}' (Source: {src})")
                        logger.info(f"  [FOUND] payment_to = {claim.payment_to} (keyword scan - payment to repairer)")
                        break

                    if "favour" in cell_text and ("repairer" in cell_text or "insured" in cell_text):
                        if "insured" in cell_text:
                            claim.payment_to = "INSURED"
                        else:
                            claim.payment_to = "REPAIRER"
                        src = f"R{r_idx+1}C{c_idx+1} ({sh_name})"
                        claim._excel_coords["payment_to"] = src
                        claim._excel_logs.append(f"  📊 payment_to: '{claim.payment_to}' (Source: {src})")
                        logger.info(f"  [FOUND] payment_to = {claim.payment_to} (keyword scan)")
                        break
                    if "favour" in cell_text:
                        for nc in range(c_idx + 1, min(c_idx + 5, len(row))):
                            next_text = " ".join(str(row[nc]).strip().lower().split())
                            if "repairer" in next_text:
                                claim.payment_to = "REPAIRER"
                                src = f"R{r_idx+1}C{nc+1} ({sh_name})"
                                claim._excel_coords["payment_to"] = src
                                claim._excel_logs.append(f"  📊 payment_to: '{claim.payment_to}' (Source: {src})")
                                break
                            elif "insured" in next_text:
                                claim.payment_to = "INSURED"
                                src = f"R{r_idx+1}C{nc+1} ({sh_name})"
                                claim._excel_coords["payment_to"] = src
                                claim._excel_logs.append(f"  📊 payment_to: '{claim.payment_to}' (Source: {src})")
                                break
                        if claim.payment_to: break
                if claim.payment_to: break
            if claim.payment_to: break

    return claim
